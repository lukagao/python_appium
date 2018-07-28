#encoding=utf-8
from appium import webdriver  
from inspect import isfunction
from openpyxl import Workbook
import aircv as ac
import datetime
import threading
from Queue import Queue
import subprocess
import sys

class BaseTest(object):
    platformName=None
    platformVersion=None
    deviceName=None
    appPackage=None
    appActivity=None
    noReset='true'
    server_url='http://localhost:4723/wd/hub'
    def __init__(self):
        self.desired_caps = {} 
        if getattr(self,'platformName'):
            self.desired_caps['platformName']=self.platformName
        else:
            raise ValueError("%s must have a platformName" % type(self).__name__)
        if getattr(self,'platformVersion'):
            self.desired_caps['platformVersion']=self.platformVersion
        else:
            raise ValueError("%s must have a platformVersion" % type(self).__name__)
        if getattr(self,'deviceName'):
            self.desired_caps['deviceName']=self.deviceName
        else:
            raise ValueError("%s must have a deviceName" % type(self).__name__)
        if getattr(self,'appPackage'):
            self.desired_caps['appPackage']=self.appPackage
        else:
            raise ValueError("%s must have a appPackage" % type(self).__name__)
        if getattr(self,'appActivity'):
            self.desired_caps['appActivity']=self.appActivity
        else:
            raise ValueError("%s must have a appActivity" % type(self).__name__)
        self.desired_caps['noReset']=self.noReset
        self.driver = webdriver.Remote(self.server_url, self.desired_caps) 
        print BaseTest.__subclasses__()
    
    def getDriver(self):
        return self.driver
    
    def getPositionByImage(self,file='shot.png'):
        mainimg='main.png'
        cv2=ac.cv2
        metlist=[cv2.TM_CCOEFF_NORMED,cv2.TM_CCORR_NORMED,cv2.TM_SQDIFF_NORMED]
        self.driver.get_screenshot_as_file(mainimg)
        mainimg=ac.imread('main.png')
        shotimg=ac.imread(file)
        for method in metlist:
            result = ac.find_all_template(mainimg, shotimg,threshold=0.5,maxcnt=1,method=method)
            if result:
                break
        if result:
            pos=result[0]
        else:
            raise ValueError('can not recognize the picture %s' %file)
        X = int(pos['result'][0])
        Y = int(pos['result'][1])
        return X,Y
    
    def clickByImage(self,file='shot.png'):
        mainimg='main.png'
        cv2=ac.cv2
        metlist=[cv2.TM_CCOEFF_NORMED,cv2.TM_CCORR_NORMED,cv2.TM_SQDIFF_NORMED]
        self.driver.get_screenshot_as_file(mainimg)
        mainimg=ac.imread('main.png')
        shotimg=ac.imread(file)
        #grayA = cv2.cvtColor(mainimg,cv2.COLOR_BGR2GRAY)
        #grayB = cv2.cvtColor(shotimg,cv2.COLOR_BGR2GRAY)
        #cv2.imshow('show',grayA)
        #cv2.waitKey(0)
        #cv2.imshow('show2',grayB)
        #cv2.waitKey(0)
        
        #result = ac.find_all_template(mainimg, shotimg,threshold=0.5,maxcnt=1)
        
        for method in metlist:
            result = ac.find_all_template(mainimg, shotimg,threshold=0.5,maxcnt=1,method=method)
            #print result
            if result:
                break
        if result:
            pos=result[0]
        else:
            raise ValueError('can not recognize the picture %s' %file)
        #img_width = mainimg.shape[1]
        #img_height = mainimg.shape[0]
        #print shotimg.shape[1],shotimg.shape[0]
        #phone_width=self.driver.get_window_size()['width']        
        #phone_height=self.driver.get_window_size()['height']
        X = int(pos['result'][0])
        Y = int(pos['result'][1])
        #print phone_width,phone_height,img_width,img_height,X,Y
        #print (phone_width/img_width)*X,(phone_height/img_height)
        self.driver.tap([(X,Y)])
    def down(self):
        print 'down'
        self.driver.quit()


def BaseRun(func):
    def inner(self):
        print 'start:'
        func(self)
    return inner


def runCMD(stdq,cmd):
    
    p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    lineNum=0
    try:
        for line in iter(p.stdout.readline,b''):
            lineNum+=1
            print line
            print '---------------------------------------------------------------------'
            if lineNum==2:
                print 'appium started'
                stdq.put('start')  
                
    except Exception as e:
        print e    
  

def runLoop(caseNum,caseList,loop,globalLoop,sheet):
    testCase=caseList[caseNum]
    tc=testCase()
    caseID=testCase.__name__
    appPackage=testCase.appPackage
    platformName=testCase.platformName
    platformVersion=testCase.platformVersion
    failTimes=0
    failReason=''
    for i in range(loop):  
        loops=i   
        for item in testCase.__dict__.values():
            if isfunction(item):
                if item.__name__!='tearDown':
                    try:
                        item(tc)
                    except Exception as e:
                        failTimes+=1
                        failReason+=str(e)
                        tc.tearDown()
                        tc.__init__()
    tc.tearDown()
    result=[caseID,appPackage,platformName,platformVersion,(loops+1)*globalLoop,failTimes,failReason]
    for x in range(len(result)):
        if result[x]==failTimes:
            originFailTimes=sheet.cell(row=caseNum+2,column=x+1).value
            if(originFailTimes):
                originFailTimes=int(originFailTimes)
            else:
                originFailTimes=0
                
            sheet.cell(row=caseNum+2,column=x+1,value=result[x]+originFailTimes)   
        else:
            sheet.cell(row=caseNum+2,column=x+1,value=result[x])      
    
def runGlobal(caseList,loop=1,globalLoop=1,multiRun=False): 
    colum=['caseID','appPackage','platformName','platformVersion','loops','failTimes','failReason']
    try:
        wb=Workbook()
        filename=datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")+'.xlsx'
        sheet=wb.active
        sheet.title='result'
        for x in range(len(colum)):
            sheet.cell(row=1,column=x+1,value=colum[x])
        for n in range(globalLoop):
            for caseNum in range(len(caseList)):        
                if(not multiRun):
                    runLoop(caseNum,caseList,loop,n+1,sheet)
                else:
                    print "start thread %s" % str(caseNum)
                    t=threading.Thread(target=runLoop,args=(caseNum,caseList,loop,n+1,sheet))
                    t.start()
                    #time.sleep(20)
                    #t.join()
            if(multiRun):
                for t in threading.enumerate():
                    if type(t)==threading.Thread:
                        t.join(60)
    except Exception as e:
        print e
    finally:  
        wb.save(filename)

def runTest(caseList,loop=1,globalLoop=1,cmds=None,multiRun=False):
    stdq=Queue()
    if cmds:
        for cmd in cmds:
            threading.Thread(target=runCMD,args=(stdq,cmd,)).start()
            while(True):
                if not stdq.empty() and stdq.get()=='start':
                    print cmd+' has been started'
                    break
    print 'start test:'
    runGlobal(caseList,loop=loop,globalLoop=globalLoop,multiRun=multiRun)  
    print 'end test:'
    #cmdprocess.terminate()
    sys.exit()
    